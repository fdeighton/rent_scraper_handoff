"""Export active comp_buildings to Excel.

Self-contained: data snapshot embedded below (taken 2026-05-28). Regenerate
the DATA constant by re-querying comp_buildings via Supabase MCP if needed.

Run from project root:
    python scraper/tools/export_comp_buildings.py
"""
from __future__ import annotations

from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

PROJECT_ROOT = Path(__file__).resolve().parents[2]

DATA = [
    # (name, address, city, scrape_url)
    ("1315 Bough Beeches Blvd", "1315 Bough Beeches Blvd", "Mississauga", "https://briarlane.ca/residential/1315-bough-beeches-boulevard"),
    ("185 Enfield", "185 Enfield Pl", "Mississauga", "https://www.gwlraresidential.com/apartments/on/mississauga/one-eighty-five/floorplans"),
    ("Avia 2", "430 Square One Dr", "Mississauga", "https://condos.ca/mississauga/avia-2-condos-465-burnhamthorpe-rd-w-430-square-one-dr"),
    ("Dialogue", "4235 Confederation Pkwy", "Mississauga", "https://bgoliving.com/residential/dialogue"),
    ("EXS Luxury", "4023 The Exchange", "Mississauga", "https://rentexs.com/floor-plans/"),
    ("Limelight Condos", "360 Square One Dr", "Mississauga", "https://condos.ca/mississauga/limelight-north-tower-360-square-one-dr-370-square-one-dr"),
    ("M City Phase 1", "3900 Confederation Pkwy", "Mississauga", "https://condos.ca/mississauga/m-city-condos-3900-confederation-pkwy-3980-confederation-pkwy"),
    ("Mirage Condominium", "339 Rathburn Rd W", "Mississauga", "https://condos.ca/mississauga/mirage-339-rathburn-rd-w"),
    ("Mississauga Place", "3620 Kaneff Crescent", "Mississauga", "https://quadrealresidential.com/mississauga/mississauga-place/floorplans/"),
    ("One Park Tower", "388 Prince of Wales Dr", "Mississauga", "https://condos.ca/mississauga/one-park-tower-388-prince-of-wales-dr"),
    ("Solstice Condos", "225 Webb Dr", "Mississauga", "https://condos.ca/mississauga/solstice-225-webb-dr"),
    ("Stak36", "395 Square One Dr", "Mississauga", "https://condos.ca/mississauga/stak36-at-condominiums-at-square-one-district-4220-living-arts-dr-395-square-one-dr"),
    ("The Arista", "3665 Arista Way", "Mississauga", "https://www.thearista.com/floorplans"),
    ("The Exchange", "4015 The Exchange", "Mississauga", "https://condos.ca/mississauga/4015-the-exchange-4015-the-exchange"),
    ("The Forestwoods", "1547 Mississauga Vlg Blvd", "Mississauga", "https://www.theforestwoods.com/floorplans"),
    ("The Huron", "2475 Hurontario St", "Mississauga", "https://2475hurontariostreet.com"),
    ("The Kay", "3610 Dixie Road", "Mississauga", "https://livingatthekay.com"),
    ("The Valleywoods", "1423 Mississauga Vlg Blvd", "Mississauga", "https://www.thevalleywoods.com/floorplans"),
    ("1 Square Phillips", "1205 Place du Frère-André", "Montreal", ""),
    ("2070 Maisonneuve", "2070 Boulevard de Maisonneuve Ouest", "Montreal", "https://www.hazelviewproperties.com/properties/2070-de-maisonneuve"),
    ("Apartments Dorchester", "1555 Boulevard René-Lévesque Ouest", "Montreal", "https://www.appartementsdorchester.com/floorplans"),
    ("Bleury", "2100 Rue de Bleury", "Montreal", "https://www.loxley.ca/communities/quartier-des-spectacles"),
    ("Brix Apartments", "1090 Rue de Bleury", "Montreal", "https://www.zumper.com/apartment-buildings/p508449/brix-mtl-quartier-ville-marie-montreal-qc"),
    ("Concorde", "375 Avenue de la Concorde", "Montreal", "https://www.loxley.ca/communities/quartier-des-spectacles"),
    ("Dorchester", "1160 Rue Saint-Mathieu", "Montreal", "https://www.maddox.ca/communities/dorchester"),
    ("Dufort Apartments", "1350 Rue du Fort", "Montreal", "https://www.capreit.ca/apartments-for-rent/montreal-qc/the-1350-dufort-apartments/"),
    ("Finale Luxury Residences", "950 Rue Saint-Antoine Ouest", "Montreal", "https://www.apartments.com/finale-luxury-residences-montreal-qc/3rbtzdd/"),
    ("Humaniti", "340 Rue de la Gauchetière Ouest", "Montreal", "https://www.humanitimontreal.com/en/rental-condos/"),
    ("La Citadelle", "2125 Rue Saint-Marc", "Montreal", "https://www.hazelviewproperties.com/properties/2125-rue-saint-marc"),
    ("Le George", "1001 Rue Lucien-L'Allier", "Montreal", "https://app.planpoint.io/le-george/le-george?lang=English"),
    ("Le Livmore", "1180 Rue de Bleury", "Montreal", "https://www.lelivmore.com/floorplans"),
    ("Le Vanguard", "2000 Rue Saint-Marc", "Montreal", "https://www.gmrre.com/residential/2000-saint-marc"),
    ("LIV Montreal", "1380 Boulevard René-Lévesque Ouest", "Montreal", "https://www.summitmanagement.ca/apartments/liv"),
    ("Maestria", "1245 Rue de Bleury", "Montreal", "https://app.planpoint.io/g/maestria-apartments/Maestria-Apartments?lang=en&customCss=true"),
    ("Mille Carre", "2061 Rue Stanley", "Montreal", "https://www.loxley.ca/communities/mille-carre"),
    ("Serra", "1124 Rue de Bleury", "Montreal", ""),
    ("The 2250 Guy Apartments", "2250 Rue Guy", "Montreal", "https://www.capreit.ca/apartments-for-rent/montreal-qc/the-2250-guy-apartments/"),
    ("VIE Apartments", "2255 Rue Saint-Mathieu", "Montreal", "https://www.irent.com/communities/2255-saint-mathieu"),
    ("103 & 105 West Lodge", "103 & 105 West Lodge Ave", "Toronto", "https://www.hazelviewproperties.com/properties/103-west-lodge"),
    ("203 Residences", "203 Jarvis St", "Toronto", "https://www.rent203residences.com/suites"),
    ("41 Dundonald", "41 Dundonald Street", "Toronto", "https://www.hazelviewproperties.com/properties/41-dundonald-street"),
    ("480 Wilson", "480 Wilson Ave", "Toronto", ""),
    ("60 Tyndall", "60 Tyndall Ave", "Toronto", "https://www.dmsproperty.com/apartments-for-rent/60-tyndall-ave"),
    ("Akoya Living", "55 Broadway Ave", "Toronto", "https://www.akoyaliving.ca/suites"),
    ("Birch House", "100 Mill St", "Toronto", "https://triconresidential.com/apartment/birch-house/#availability"),
    ("Cherry House", "125 Mill St", "Toronto", "https://triconresidential.com/apartment/cherry-house/#availability"),
    ("College West", "897 College St", "Toronto", "https://www.collegewest.com/suites"),
    ("Corner on Broadway", "223 Redpath Ave", "Toronto", "https://thecornerrentals.com/suites"),
    ("Dufferin Crossing", "295 Dufferin Street", "Toronto", "https://www.295dufferinstreet.com/suites"),
    ("E18HTEEN", "18 Erskine Ave", "Toronto", "https://www.myrental.ca/apartments-for-rent/18-erskine-ave"),
    ("eCentral", "15 Roehampton Ave", "Toronto", "https://www.ecentralliving.com/rental-suites"),
    ("Elm Ledbury", "25 Dalhousie Street", "Toronto", "https://www.elmledbury.ca/suites"),
    ("Evolv", "25 Nicholas Avenue", "Toronto", "https://www.evolvrentals.com/"),
    ("Four Fifty", "450 Front St", "Toronto", "https://livingatfourfifty.com/suites/"),
    ("Galleria 01 & 02", "1245 Dupont St", "Toronto", ""),
    ("Huntley Apartments", "77 Huntley Street", "Toronto", "https://www.rpmsi.ca/property/huntley-apartments-77-huntley/"),
    ("IMMIX", "7 Grosvenor Street", "Toronto", "https://quadrealresidential.com/toronto/immix/floorplans/"),
    ("Imperial Towers", "90 Tyndall Ave", "Toronto", "https://rentals.ca/toronto/90-tyndall-avenue"),
    ("Jazz", "167 Church Street", "Toronto", "https://rent.concertproperties.com/toronto/jazz/floorplans"),
    ("Lakeview Towers", "22 Close Ave", "Toronto", "https://amelinproperties.com/apartments/lakeview-towers"),
    ("Litho", "740 Dupont St", "Toronto", "https://www.litholiving.com/floorplans"),
    ("Livmore", "55 Gerrard Street East", "Toronto", "https://www.gwlraresidential.com/apartments/on/toronto/the-livmore-bay-gerrard/floorplans"),
    ("LNX Residences", "2376 Dundas St W", "Toronto", "https://www.lnxrentals.com/suites"),
    ("Lord Dufferin Apartments", "200 Dufferin St", "Toronto", "https://www.hazelviewproperties.com/properties/200-dufferin-street"),
    ("Maple House", "397 Front St E", "Toronto", "https://triconresidential.com/apartment/maple-house/#availability"),
    ("Marlow (Block D)", "980 Dufferin St", "Toronto", ""),
    ("Mirvish Village", "748 Bathurst St", "Toronto", "https://mirvish-residential-rentcafewebsite.securecafe.com/onlineleasing/mirvish-residential/floorplans"),
    ("Parker", "200 Redpath Avenue", "Toronto", "https://www.parkerlife.ca/suites"),
    ("Pivot", "35 Green Field Ave", "Toronto", "https://www.livingatpivot.com/floorplans"),
    ("Sherbourne", "191 Sherbourne Street", "Toronto", "https://www.maddox.ca/communities/cabbagetown"),
    ("Shorecrest", "35 Spencer Ave", "Toronto", "https://rentals.ca/toronto/shorecrest"),
    ("Sixty Five Broadway", "65 Broadway Ave", "Toronto", "https://docs.google.com/spreadsheets/d/18sfDzaCgFPrWCeUmawrBylkR9Fgkr4o8CN86AKIeTAI/export?format=csv&gid=0"),
    ("Sloane - Tower A/B", "3450 Dufferin Ave", "Toronto", "https://www.sloanelife.ca/suites"),
    ("Sloane - Tower C", "3450 Dufferin Ave", "Toronto", "https://www.sloanelife.ca/suites"),
    ("Spencer Apartments", "47 Spencer Ave", "Toronto", "https://www.dmsproperty.com/apartments-for-rent/spencer-apartments"),
    ("Station House", "1141 Bloor St W", "Toronto", ""),
    ("Story of Midtown", "75 Broadway Ave", "Toronto", "https://www.mystorymidtown.com/suites"),
    ("Strada", "555 College St", "Toronto", "https://livingatstrada.capreit.ca/suites"),
    ("The Brixton", "400 Dufferin St", "Toronto", "https://www.rentbrixton.ca/floorplans"),
    ("The Campbell", "299 Campbell Ave", "Toronto", "https://www.rentthecampbell.com/suites"),
    ("The Diamond", "980 Lansdowne Ave", "Toronto", "https://www.diamondrentals.ca/suites"),
    ("The Hampton", "101 Roehampton Ave", "Toronto", "https://thehampton.ca/floorplans/"),
    ("The Ivy", "8 Gloucester Street", "Toronto", "https://triconresidential.com/apartment/the-ivy/"),
    ("The Maples", "20 Carlton Street", "Toronto", "https://www.themaplescomplex.ca/"),
    ("The Montgomery", "2388 Yonge St", "Toronto", "https://www.themontgomery.ca/floorplans"),
    ("The Selby", "25 Selby St", "Toronto", "https://triconresidential.com/apartment/the-selby/"),
    ("The Spoke", "485 Perth Ave", "Toronto", ""),
    ("The Taylor", "57 Spadina Avenue", "Toronto", "https://triconresidential.com/apartment/the-taylor/"),
    ("The Whitney", "71 Redpath Ave", "Toronto", "https://www.thewhitneyonredpath.com/apartments"),
    ("Tippett Park Two", "18 Tretti Way", "Toronto", "https://www.shiplake.com/tippett-park"),
    ("Trilogy on King", "1100 King St W", "Toronto", "https://www.trilogyliving.ca/floorplans"),
    ("Tyndall", "115 & 135 Tyndall Ave", "Toronto", "https://www.maddox.ca/communities/tyndall"),
    ("Waldorf Tower", "1251 King St W", "Toronto", "https://www.1251king.ca/"),
    ("Waverley", "484 Spadina Ave", "Toronto", "https://www.waverleylife.ca/floorplans"),
    ("Wellesley Apartments", "100 Wellesley Street E", "Toronto", "https://www.capreit.ca/apartments-for-rent/toronto-on/wellesley-apartments/"),
    ("Wellesley Square", "155 Wellesley Street", "Toronto", "https://www.155wellesleystreet.com/suites"),
    ("West House", "88 Bathurst St", "Toronto", "https://sightmap.com/embed/1ywym6e1wq0"),
    ("XO Condos", "270 Dufferin St", "Toronto", "https://condos.ca/toronto/xo-condos-270-dufferin-st"),
    ("19 on the Greenway", "3619-3681 Arbutus St", "Vancouver", "https://19onthegreenway.com/floorplans"),
    ("Archetype Residences", "220 E 1st Ave", "Vancouver", "https://quadrealresidential.com/vancouver/archetype-residences/floorplans/"),
    ("Chroma", "1888 Scotia St", "Vancouver", "https://chromavancouver.ca/floorplans.html"),
    ("English Bay Residences", "1668 Davie St", "Vancouver", "https://englishbayresidences.com"),
    ("Revolve", "3572 Glen Dr", "Vancouver", "https://revolve.petersonrentals.com/floorplans"),
    ("Senakw I", "1450 Pennyfarthing Dr", "Vancouver", ""),
    ("Signal", "8428 Ash St", "Vancouver", "https://www.rentsignal.ca/suites"),
    ("The Pendrell", "1108 Pendrell St", "Vancouver", "https://thependrell1770.capreit.ca/suites"),
    ("The Stories", "2488 Granville St", "Vancouver", "https://thestoriessouthgranville.com/floorplans"),
    ("Zephyr Apartments", "1661 Davie St", "Vancouver", "https://devonproperties.com/properties/residential/zephyr-apartments/"),
    ("The Millway", "185 Millway", "Vaughan", "https://themillway.ca/floor-plans/"),
    ("Daniels on Parliament", "131 Parliament St, Toronto, ON", "(Unspecified)", "https://www.danielsonparliament.com/floorplans/"),
]


def main() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Comp Buildings"

    header_fill = PatternFill(start_color="1B2A4A", end_color="1B2A4A", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    city_fill = PatternFill(start_color="C0392B", end_color="C0392B", fill_type="solid")
    city_font = Font(bold=True, color="FFFFFF", size=12)

    headers = ["Name", "Address", "City", "Scrape URL"]
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 22

    current_city = object()
    for name, address, city, url in DATA:
        if city != current_city:
            ws.append([city, "", "", ""])
            r = ws.max_row
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
            sc = ws.cell(row=r, column=1)
            sc.fill = city_fill
            sc.font = city_font
            sc.alignment = Alignment(horizontal="left", vertical="center")
            ws.row_dimensions[r].height = 20
            current_city = city
        ws.append([name, address, city, url])

    ws.freeze_panes = "A2"
    for col_idx, width in {1: 36, 2: 42, 3: 16, 4: 80}.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    stamp = date.today().isoformat()
    out_path = PROJECT_ROOT / f"comp_buildings_export_{stamp}.xlsx"
    wb.save(out_path)
    print(f"Wrote {len(DATA)} active buildings to:")
    print(f"  {out_path}")


if __name__ == "__main__":
    main()
